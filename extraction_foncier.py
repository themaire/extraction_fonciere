# V2.4
# Debug export fichier excel ( lignes vides dans le cas de lot de propriétaires )
# Ajout du module m_bdd
# Ajout de la differentiation du type de machine

import sys
import os

from dotenv import load_dotenv
load_dotenv()

from datetime import date

import openpyxl
from openpyxl.styles.alignment import Alignment
import subprocess
import shutil
from time import sleep

from utils.m_bdd import *

# Charger le fichier .env (ajuster le chemin si nécessaire)
load_dotenv()

pc_type = os.name

# Récupérer le répertoire courant du script
current_dir = os.path.dirname(os.path.abspath(__file__))
GABARIT_EXCEL = os.getenv('GABARIT_EXCEL')
GABARIT_EXCEL = os.path.join(current_dir, GABARIT_EXCEL)

BASE_DIR = os.getenv('BASE_DIR')
TABLEPARCELLES = os.getenv('TABLEPARCELLES')
DBHOST = os.getenv('DBHOST')
DBUSER = os.getenv('DBUSER')
DBPASSWD = os.getenv('DBPASSWD')

DBUSER_HISTO = os.getenv('DBUSER_HISTO')
DBHOST_HISTO = os.getenv('DBHOST_HISTO')
DBPASSWD_HISTO = os.getenv('DBPASSWD_HISTO')
BDD_HISTO = os.getenv('BDD_HISTO')
TABLE_HISTO = os.getenv('TABLE_HISTO')

prog_pgsql2shp = os.getenv('prog_pgsql2shp')

INPUT_DIR = os.path.join(BASE_DIR, "a_traiter")
DONE_DIR = os.path.join(INPUT_DIR, "fait")
RESULT_DIR = os.path.join(BASE_DIR, "resultats")
print('INPUT_DIR', INPUT_DIR)
print('DONE_DIR', DONE_DIR)
print('RESULT_DIR', RESULT_DIR)

DIRS = {'BASE_DIR' :BASE_DIR,'INPUT_DIR' : INPUT_DIR,'DONE_DIR' : DONE_DIR,'RESULT_DIR' : RESULT_DIR}
COLUMS = 5

# Test si on a bien mis un fichier en premier paramètre.
# Exit si ce dernier n'en est pas un ou pas de paramètre.
if len(sys.argv) > 1:
    write_histo = str(sys.argv[2])
    print("write_histo :", write_histo)

    path_file = str(sys.argv[1])
    current_file = os.path.join(INPUT_DIR, path_file.split("/")[-1]) # Mode normal des choses
    print("current_file :", current_file)
    print()

    #current_file = os.path.join(DONE_DIR, path_file.split("\\")[-1]) # En mode création de l'historique
    
    if os.path.isfile(current_file):
        # print('current_file' , 'is file.')
        print()
    else:
        print("Is not a file...")
        print("path_file: ", path_file)
        print("...")
        print("current_file :", current_file)
        exit()
else:
    print("Pas de fichier en entrée à traiter.")
    exit(1)

# Verifie les dossiers de travail
def checkDirs(DIRS):
    '''
    Test si tous les dossiers existent avant de commencer quoi que ce soit.
    '''
    check = True
    for i in DIRS:
        if not os.path.isdir(DIRS[i]):
            print(i , ' nor a dir'':', DIRS[i])
            check = False
    return False

# Feuille excel des parcelles EN SORTIE
def query_to_xlsx(query_name, feuille):
    lig = 3
    col = 1
    for ligne in query_name :
        # Pour chaque ligne de résultat
        for cellule in ligne:
            # Pour chaque colonne dans une ligne
            idx = ligne.index(cellule)
            if idx == 0:
                cell = feuille.cell(row=lig, column=col)
                cell.value = cellule
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                col += 1
            elif idx <= 8:
                # Si on est dans les colonnes VERTES (CARACTERISTIQUES PARCELLES)
                firstCell = feuille.cell(row=lig, column=1)
                ligref = lig - 1
                if firstCell.value != feuille.cell(row=ligref, column=1).value:
                    # Si la cellule de la premiere colonne (idx 1) est differente de celle juste au dessus 
                    cell = feuille.cell(row=lig, column=col)
                    cell.value = cellule
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    col += 1
                else:
                    cell = feuille.cell(row=lig, column=col)
                    cell.value = '"'
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    col += 1
            elif idx == 9:
                # Si on est dans la colonne masque (lettre J), on affiche la valeur vide definie dans la requete (champ sep='').
                cell = feuille.cell(row=lig, column=col)
                cell.value = cellule
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                col += 1
            elif idx >= 10 and idx < 23:
                # firstCell = feuille.cell(row=lig, column=1)
                # cellpro = feuille.cell(row=lig, column=10)
                # ligref = lig - 1
                # if (cellpro.value == feuille.cell(row=ligref, column=10).value) and (firstCell.value == feuille.cell(row=ligref, column=1).value):
                #     cell = feuille.cell(row=lig, column=col)
                #     cell.value = ''
                #     cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                #     col += 1
                # else:
                cell = feuille.cell(row=lig, column=col)
                cell.value = cellule
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                col += 1
            else:
                cell = feuille.cell(row=lig, column=col)
                cell.value = cellule
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                col += 1
        lig += 1
        col = 1
        
# Feuille Excel des propriétaires EN SORTIE
def query_to_xlsx_2(query_name, feuille):
    lig = 3
    col = 1
    l3 = ''
    for ligne in query_name :
        # Pour chaque ligne de résultat
        for cellule in ligne:
            # Pour chaque colonne dans une ligne
            idx = ligne.index(cellule)
            if idx == 0:
                cell = feuille.cell(row=lig, column=col)
                cell.value = cellule
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                col += 1
            elif idx <= 7:
                # Si on est dans les colonnes SAUMON (CARACTERISTIQUES PROPRIETAIRES)
                cellpro = feuille.cell(row=lig, column=1)
                ligref = lig - 1
                if cellpro.value != feuille.cell(row=ligref, column=1).value:
                    cell = feuille.cell(row=lig, column=col)
                    cell.value = cellule
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    col += 1
                else:
                    cell = feuille.cell(row=lig, column=col)
                    cell.value = ''
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    col += 1
            elif idx == 8:
                cellpro = feuille.cell(row=lig, column=1)
                ligref = lig - 1
                if cellpro.value != feuille.cell(row=ligref, column=1).value:
                    cell = feuille.cell(row=lig, column=col)
                    cell.value = cellule
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    col += 1
                else:
                    ligdest = lig - 1
                    celldest = feuille.cell(row=ligdest, column=9)
                    cell_val = celldest.value
                    celldest.value = f"{cell_val}, {cellule}"
            elif idx == 9:
                if l3 != '':
                    cellpro = feuille.cell(row=lig, column=1)
                    ligref = lig - 1
                    if cellpro.value != feuille.cell(row=ligref, column=1).value:
                        celldest = feuille.cell(row=lig, column=9)
                        cell_val = celldest.value
                        celldest.value = f"{cell_val} LOT {cellule}"
                    else:
                        ligdest = lig - 1
                        celldest = feuille.cell(row=ligdest, column=9)
                        cell_val = celldest.value
                        celldest.value = f"{cell_val} LOT {cellule}"
        cellpro = feuille.cell(row=lig, column=1)
        ligref = lig - 1
        if cellpro.value != feuille.cell(row=ligref, column=1).value:
            lig += 1
            col = 1
        else:
            col = 1

# S'occupe de remplir le fichier excel EN SORTIE à partir d'une requete
def writeExcel(query, wb, workbook_name, connection, type = 1):
            try:
                result = db_query(connection, query)
                wb = wb[workbook_name]
                if(type == 1):
                    query_to_xlsx(result, wb)
                elif(type == 2):
                    query_to_xlsx_2(result, wb)
            except:
                print("c'est mal passée pour ", workbook_name, ":" ,query)
                return 1
            connection.close()
            return 0

# récupere la liste des fichiers à traiter (pour la liste de choix) NON UTILISE POUR LE MOMENT - Gardé au cas où
def files_to_do(input, mode = 'name'):
    list = []
    files = os.scandir(input)
    for i in files:
        if i.is_file():
            if(mode == 'name'):
                list.append(i.name)
            elif(mode == 'path'):
                list.append(i.path)
            else:
                list.append(i.name)
    return list

# récupere les infos contenues dans le nom du fichier Excel EN ENTREE
def extraction_name(filename):
    #print("extraction_name() filename :", str(filename))
    
    if (pc_type == 'nt'):
        if(filename[-5:] == ".xlsx"):
            infos = filename.split('\\')[-1].split('_Modele')[0]
        else:
            infos = filename.split('_Modele')[0].split('/')[-1]
    elif (pc_type == 'posix'):
        infos = filename.split('\\')[-1].split('_Modele')[0]
    
    ssplit = infos.split('_')

    extract_infos = { "all" : infos, "initiales" :ssplit[0], "code" : ssplit[1], "dep" : ssplit[2], "description" : ssplit[3]}
    return extract_infos

def open_excel(filename):
    """
    Opens an Excel file and returns the rows of data from the active sheet.

    Args:
    filename (str): The name of the Excel file to be opened.

    Returns:
    data (iterable): Rows of data from the active sheet of the Excel file.
    """
    xlsx = openpyxl.load_workbook(filename)
    sheet = xlsx.active
    data = sheet.rows
    return data

# fabrique une liste python à partir d'un fichier excel
def xls2list(excel_filename):
    '''
    Retourne une simple liste python refletant le contenu du fichier excel.
    '''
    parcellsList = []
    
    try:
        data = open_excel(str(excel_filename))
    except:
        print('Le fichier EXCEL a un souci de lecture-ecriture.')
        return 1

    for i in data:
        if(i == 0):
            pass

        subList = []
        for j in range(COLUMS):
            cellvalue = str(i[j].value)
            
            # print(cellvalue)
            
            subList.append(cellvalue)
        parcellsList.append(subList)

    print("Il y a", str(len(parcellsList)), 'parcelles dans le fichier.')
    return parcellsList

# 2 fonctions pour les requetes
def wheremakR(list_parcelles):
    '''
    Fabrique le where de la requete qui sert à filtrer les parcelles
    de la demande au format fichier Excel.
    '''

    where = "WHERE code_parcelle in ("
    cpt = 0

    # On ne prend pas le premier element de la liste car il s'agit 
    # de l'entete des colonnes du fichier excel
    for i in list_parcelles:
        # print("cpt = ", cpt, "data :", i)

        if (cpt > 0):
            where += ", "
        
        cookedCodeParcelle = "'" + f"{i[1]:0>5}" + f"{i[2]:0>3}" + f"{i[3]:0>2}" + f"{i[4]:0>4}" + "'"

        # where += "(insee = '{}' AND section = '{}' AND numero = '{}')".format(i[1], i[3], i[4])
        #where += "code_parcelle = lpad(" + i[1] + ",5, '0') || lpad(" + i[2] + ", 4, '0') || lpad(" + i[3] + ", 2, '0') || lpad(" + i[4] + ", 4, '0')"
        where += cookedCodeParcelle
        
        cpt += 1

    return where + ')'

def querymakR(where):
    '''
    Fabrique la requette pour l'export shapefile de l'extraction.
    '''
    query = "SELECT code_parcelle, insee, commune, lieu_dit, prefix, section, numero, surface_ha, culture_dominante, date_acte, type_filiation, type_proprietaire, division_lot, nombre_lots, geompar FROM liste_parcelles "

    query += where
    return query

# lance le programme pgsql2shp
def progCommand(directory, infos, where):
    '''
    Fabrique la commande a executer pour appeler pgsql2shp.
    '''
    prog_query = querymakR(where)
    # output = os.path.join(RESULT_DIR , infos)
    output = os.path.join(RESULT_DIR , directory, infos)
    print("output folder: ", output)
    #output = "../resultats" + infos

    prog_options = '-f "{}" -h '.format(output) + DBHOST + ' -u "' + DBUSER + '" -P "' + DBPASSWD + '" foncier "{}"'.format(prog_query)
    # print(prog_options)

    prog_command = r'{} {}'.format(prog_pgsql2shp, prog_options)
    # print()
    print("programm :", prog_command)
    # print()
    return prog_command

###
##### Debut du script
###

def main(writeFiles = True, writeHisto = False):
    try:
        checkDirs(DIRS)
    except:
        exit()

    # Extraire les infos essentielles de l'extraction
    # exemple : MT_23-5195A_Blancs-Coteaux_Vertus
    print("current file :", current_file)
    extraction = extraction_name(current_file)
    extraction_infos = extraction['all']
    print("extraction_infos :", extraction_infos)

    # Lire le fichier Excel passé en paramètre et le transformer en liste python
    listParcells = xls2list(current_file)
    print()
    print("4 premieres listparcelles ", listParcells[:4], " ... ...")

    # Tres important pour les 3 requetes qui vont suivre
    where = wheremakR(listParcells)
    print()
    #print("---> where ", where)

    if (writeFiles):
        # Si on veut ecrire les fichier shape et excel :
        print("Démarrage du traitement des fichiers de sortie :")
        print()

        # Création du dossier de destination
        date_infos = str(str( date.today() ) + "_" + extraction_infos)

        try:
            os.makedirs(RESULT_DIR)
            print("Dossier", RESULT_DIR, "créé.")
            print()
        except FileExistsError:
            print('RESULT_DIR', RESULT_DIR)
            pass
        print( 'directory_path = os.path.join(RESULT_DIR, date_infos)' )
        print("date_infos :", date_infos)
        directory_path = os.path.join(RESULT_DIR, date_infos)
        print("directory_path :", directory_path)

        try:
            os.makedirs(directory_path)
            print("Dossier", directory_path, "créé.")
            print()
        except FileExistsError:
            print("Le dossier de resultat existe deja.")
            print()
            pass

        # Export en shape des parcelles
        subprocess.run(progCommand(directory_path, date_infos, where), shell = True)

        # Création du fichier Excel
        excel_file = os.path.join(directory_path, extraction_infos + ".xlsx")

        shutil.copy(GABARIT_EXCEL, excel_file) # On rapatrie le modèle vide dans le nouveau fichier cible à remplir
        sleep(.1)
        wb = openpyxl.load_workbook(excel_file) # Chargement du fichier Excel

        ## Remplissage du fichier Excel
        try:
            ## 1er onglet - Partie 1
            connection_local = db_connect(DBHOST, DBUSER, DBPASSWD, "foncier")
            parcellQuery = "SELECT code_parcelle, insee, commune, lieu_dit, prefix, section, numero, surface_ha, culture_dominante, '' as foo, date_acte, type_filiation, type_proprietaire, detail_droit, nom, categorie, detail_categorie, adresse, date_naissance, adresse_naissance, siren, division_lot, nombre_lots, code_lot, superficie_lot, detail_droit_lot, nom_lot, categorie_lot, detail_categorie_lot, adresse_lot, date_naissance_lot, adresse_naissance_lot, siren_lot, idprocpte_lot FROM " + TABLEPARCELLES + " {};".format(where)
            # print(parcellQuery)
            writeExcel(parcellQuery, wb, 'Extraction par parcelles', connection_local, 1)
            
            ## 2eme onglet
            connection_local = db_connect(DBHOST, DBUSER, DBPASSWD, "foncier")
            personQuery = "SELECT code_parcelle, nom, categorie, detail_categorie, adresse, date_naissance, adresse_naissance, siren, code_parcelle FROM proprietaires_par_parcelles" + " {};".format(where)
            writeExcel(personQuery, wb, 'Extraction par proprietaires', connection_local, 2)
            
            wb.save(excel_file)
            wb.close()

            ## Déplacer le fichier Excel venant d'etre utilisé
            try:
                os.makedirs(DONE_DIR)
                print("Dossier", DONE_DIR, "créé.")
                print()
            except:
                pass
            moved_excel = os.path.join(DONE_DIR, extraction_infos + ".xlsx")
            shutil.move(path_file, moved_excel)
        except Exception as e:
            print(f"Une erreur s'est produite : {e}")

    else:
        print("NE PAS écrire les fichiers")
        print()

    ## Enregistrement dans la table d'historique
    if (writeHisto == '1'):
        # Récuperation des codes parcelles concernés dans le PC portable CENCA
        connection_local = db_connect(DBHOST, "postgres", DBPASSWD, "foncier")

        # Tout vient de la variable 'where'
        codParcellQuery = "SELECT DISTINCT code_parcelle FROM " + TABLEPARCELLES + " {};".format(where)
        print("codParcellQuery :", codParcellQuery)
        codParcell_result = db_query(connection_local, codParcellQuery)
        connection_local.close()

        # Création de la requete permettant d'INSERT dans la table d'histo foncière
        inserts = ""
        for i in codParcell_result:
            # print("i :", i)
            inserts += "('" + str(i)[2:][:-3] + "', '" + extraction['description'] + "', '" + extraction['initiales'] + "', '" + str(extraction['dep']) + "'), "
        inserts_histo = inserts[:-2]
        histoQuery = "INSERT INTO " + TABLE_HISTO + "(hfo_code_parcelle, hfo_info_extraction, hfo_demandeur, hfo_dep) VALUES {};".format(inserts_histo)
        print(histoQuery)
        print()

        connection_histo = db_connect(DBHOST_HISTO, DBUSER_HISTO, DBPASSWD_HISTO, BDD_HISTO)
        connection_histo.autocommit = True
        lizmap_result = db_query(connection_histo, histoQuery, type = 'insert')
        print(lizmap_result)
        connection_histo.close()
    else:
        print("NE PAS écrire dans la table d'historique.")
        print()
    
    exit(0)

main(writeFiles = True, writeHisto = write_histo) # writeFiles = True par defaut **** --- Si on veut ecrire les fichier shape et Excel